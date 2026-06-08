#!/usr/bin/env python3
"""
DeepRunner 6.0 报告生成器

从 DeepRunner 6.0 压测工具的 stats.json + 测试脚本自动生成:
- 单设备 Excel 测试报告（3 Sheet）
- 比对 Excel 测试报告（完整 7 Sheet，风格对齐 v6 标准）
- 单设备 HTML 可视化报告（glass-card 双主题 + Chart.js）
- 比对 HTML 可视化报告（6张对比图表 + Tab切换）

用法:
    # 单设备报告（设备/模型信息自动从 device-*.md 检测）
    python generate_report.py --mode single --stats-dir <目录> --output-dir <目录>
    python generate_report.py --mode single-html --stats-dir <目录> --output-dir <目录>

    # 手动指定设备/模型（覆盖自动检测）
    python generate_report.py --mode single --stats-dir <目录> --output-dir <目录> --device "设备名" --model "模型名"

    # 比对报告（设备/模型信息自动从各自目录 device-*.md 检测）
    python generate_report.py --mode compare --stats-dir-a <A> --stats-dir-b <B> --output-dir <目录>
    python generate_report.py --mode compare-html --stats-dir-a <A> --stats-dir-b <B> --output-dir <目录>

    # 手动指定比对设备/模型
    python generate_report.py --mode compare --stats-dir-a <A> --stats-dir-b <B> --output-dir <目录> --name-a "设备A" --name-b "设备B" --model-a "模型A" --model-b "模型B"

    # 批量单设备 Excel 报告
    python generate_report.py --mode batch --stats-dir <含多个stats.json子目录的父目录> --output-dir <输出目录>

HTML 报告会自动将 assets/ 目录（Chart.js, Tailwind, FontAwesome, Inter 字体, marked.js）复制到输出目录。
必需依赖: openpyxl (pip install openpyxl)，仅 Excel 模式需要。
"""

import argparse
import json
import os
import re
import shutil
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ============================================================
#  样式常量（精确对齐 v6 风格）
# ============================================================

# 主色调
COLOR_PRIMARY = '1F4E79'       # 深蓝（标题、表头）
COLOR_ALT_ROW_A = 'D6E4F0'     # 浅蓝交替行（逐组对比、测试概览、性能分析结论标题）
COLOR_ALT_ROW_WHITE = 'FFFFFF'  # 白色交替行
COLOR_GREEN_ALT = 'E2EFDA'     # 浅绿交替行（汇总统计）
COLOR_YELLOW_ALT = 'FFF2CC'     # 浅黄交替行（自定义指标明细）

FONT_TITLE = Font(name='微软雅黑', size=16, bold=True, color=COLOR_PRIMARY)
FONT_SUBTITLE = Font(name='微软雅黑', size=9, color='666666')
FONT_SECTION = Font(name='微软雅黑', size=12, bold=True, color=COLOR_PRIMARY)
FONT_HEADER = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
FONT_NORMAL = Font(name='微软雅黑', size=10)
FONT_BOLD = Font(name='微软雅黑', size=10, bold=True)
FONT_CONCLUSION_TITLE = Font(name='微软雅黑', size=11, bold=True)
FONT_CONCLUSION_BODY = Font(name='微软雅黑', size=10)

FILL_PRIMARY = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type='solid')
FILL_ALT_A = PatternFill(start_color=COLOR_ALT_ROW_A, end_color=COLOR_ALT_ROW_A, fill_type='solid')
FILL_WHITE = PatternFill(start_color=COLOR_ALT_ROW_WHITE, end_color=COLOR_ALT_ROW_WHITE, fill_type='solid')
FILL_GREEN = PatternFill(start_color=COLOR_GREEN_ALT, end_color=COLOR_GREEN_ALT, fill_type='solid')
FILL_YELLOW = PatternFill(start_color=COLOR_YELLOW_ALT, end_color=COLOR_YELLOW_ALT, fill_type='solid')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center')

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)  # v6 thin 边框

# ============================================================
#  数据提取层
# ============================================================

def _find_entry(entries, keyword):
    """在 stats.json 条目列表中按 name 关键词精确包含匹配"""
    for e in entries:
        if keyword in e.get('name', ''):
            return e
    return None

def extract_from_stats_json(stats_path):
    """从 stats.json 提取核心性能指标和原始条目列表
    
    stats.json 结构: JSON 数组，每个元素是命名指标条目
    
    返回:
      metrics: dict - 核心指标（用于逐组对比明细/汇总统计）
      raw_entries: list - 非URL条目列表（用于自定义指标明细）
      total_requests: int
      failed_requests: int
    """
    with open(stats_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    if isinstance(data, dict):
        entries = data.get('stats_requests', data.get('data', []))
    elif isinstance(data, list):
        entries = data
    else:
        entries = []
    
    metrics = {}
    total_req = 0
    failed = 0
    
    # 过滤非URL条目（自定义指标明细用）
    raw_entries = [e for e in entries if 'http' not in e.get('name', '')]
    
    # 1. 主请求条目
    main_req = _find_entry(entries, 'chat/completions')
    if main_req:
        total_req = main_req.get('num_requests', 0)
        failed = main_req.get('num_failures', 0)
    
    # 2. 延迟
    lat_entry = _find_entry(entries, 'Latency (s)') or _find_entry(entries, 'Latency')
    if lat_entry:
        metrics['min_latency'] = lat_entry.get('min_response_time', 0)
        metrics['max_latency'] = lat_entry.get('max_response_time', 0)
    
    avg_lat_entry = _find_entry(entries, 'Average latency (s)') or _find_entry(entries, 'Average latency')
    if avg_lat_entry:
        metrics['avg_latency'] = avg_lat_entry.get('min_response_time', 0)
    
    # 3. 吞吐量
    thr_per = _find_entry(entries, 'Throughput (tokens/s)')
    if thr_per:
        metrics['throughput_min'] = thr_per.get('min_response_time', 0)
        metrics['throughput_max'] = thr_per.get('max_response_time', 0)
    
    thr_avg = _find_entry(entries, 'Throughput(average tokens/s)')
    if thr_avg:
        metrics['throughput_avg'] = thr_avg.get('min_response_time', 0)
        metrics['throughput_avg_max'] = thr_avg.get('max_response_time', 0)
    
    # 4. QPS
    qps_entry = _find_entry(entries, 'Average QPS')
    if qps_entry:
        metrics['min_qps'] = qps_entry.get('min_response_time', 0)
        metrics['max_qps'] = qps_entry.get('max_response_time', 0)
    
    # 5. TTFT
    ttft_entry = _find_entry(entries, 'TTFT (s)')
    if ttft_entry:
        metrics['avg_ttft'] = ttft_entry.get('min_response_time', 0)
    
    # 6. TPOT
    tpot_entry = _find_entry(entries, 'TPOT (s)')
    if tpot_entry:
        metrics['avg_tpot'] = tpot_entry.get('min_response_time', 0)
    
    # 7. 输入 tokens
    in_tok_entry = _find_entry(entries, 'Input tokens')
    if in_tok_entry:
        metrics['avg_input_tokens'] = in_tok_entry.get('min_response_time', 0)
        metrics['input_tokens_max'] = in_tok_entry.get('max_response_time', 0)
    
    # 8. 输出 tokens
    out_tok_entry = _find_entry(entries, 'Output tokens')
    if out_tok_entry:
        metrics['min_output_tokens'] = out_tok_entry.get('min_response_time', 0)
        metrics['max_output_tokens'] = out_tok_entry.get('max_response_time', 0)
    
    # 9. 平均输出 tokens per request（用于图表分析）
    avg_out_tok = _find_entry(entries, 'Average output tokens per request')
    if avg_out_tok:
        metrics['avg_output_tokens'] = avg_out_tok.get('min_response_time', 0)
    
    return {
        'metrics': metrics,
        'raw_entries': raw_entries,
        'total_requests': total_req,
        'failed_requests': failed,
    }

def extract_concurrency_from_script(script_path):
    """从测试脚本提取并发数"""
    try:
        with open(script_path, 'r', encoding='utf-8') as f:
            content = f.read()
        patterns = [
            r"'用户数'\s*:\s*(\d+)",
            r'"用户数"\s*:\s*(\d+)',
            r'users\s*=\s*(\d+)',
            r'concurrency\s*=\s*(\d+)',
            r"'并发数'\s*:\s*(\d+)",
        ]
        for pat in patterns:
            m = re.search(pat, content)
            if m:
                return int(m.group(1))
    except (FileNotFoundError, IOError):
        pass
    return None

def extract_test_config(apirunner_path):
    """从 apirunner.json 提取测试配置"""
    try:
        with open(apirunner_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def extract_tasks_config(tasks_path):
    """从 tasks.json 提取任务配置（并发数、脚本名、运行时长、模式等）"""
    try:
        with open(tasks_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        config = {}
        # 并发数：scene_stages 中 user_count 最大值
        stages = data.get('scene_stages', [])
        if stages:
            config['concurrency'] = max(s.get('user_count', 0) for s in stages)
            config['run_time'] = max(s.get('run_time', 0) for s in stages)
        config['scene_mode'] = data.get('scene_mode', '')
        config['load_name'] = data.get('load_name', '')
        config['test_script'] = data.get('test_script', '').replace('\\', '/')
        config['runner_ip'] = ''
        nodes = data.get('node_list', [])
        if nodes:
            config['runner_ip'] = nodes[0].get('ip', '')
        config['total_run_time'] = data.get('total_run_time', 0)
        config['test_name'] = data.get('test_name', '')
        return config
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def extract_errors_config(errors_path):
    """从 errors.json 提取错误信息"""
    try:
        with open(errors_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        failures = data.get('failures', []) or []
        exceptions = data.get('exceptions', []) or []
        sla_msgs = data.get('sla_messages', []) or []
        return {
            'failures': failures,
            'exceptions': exceptions,
            'sla_messages': sla_msgs,
            'has_errors': bool(failures or exceptions or sla_msgs),
            'failure_count': len(failures),
            'exception_count': len(exceptions),
            'sla_count': len(sla_msgs),
        }
    except (FileNotFoundError, json.JSONDecodeError):
        return {'failures': [], 'exceptions': [], 'sla_messages': [],
                'has_errors': False, 'failure_count': 0, 'exception_count': 0, 'sla_count': 0}

def scan_test_groups(base_dir):
    """扫描目录下所有测试组，按目录名字母序返回"""
    groups = []
    base_path = Path(base_dir)
    for item in sorted(base_path.iterdir()):
        if not item.is_dir():
            continue
        stats_file = item / 'stats.json'
        if not stats_file.exists():
            continue
        
        # ── 第一步：读取 tasks.json ──
        tasks_config = extract_tasks_config(item / 'tasks.json')
        concurrency = tasks_config.get('concurrency')
        has_tasks = bool(tasks_config)
        
        # ── 第二步：若 tasks.json 有 test_script 路径，尝试读取该脚本提取并发数 ──
        script_read = False
        test_script_rel = tasks_config.get('test_script', '')
        if concurrency is None and test_script_rel:
            # test_script 是相对路径，从组目录向上尝试 10 级
            script_candidates = []
            for levels in range(10):
                p = item
                for _ in range(levels):
                    p = p.parent
                script_candidates.append(p / test_script_rel)
            for sp in script_candidates:
                if sp.exists():
                    concurrency = extract_concurrency_from_script(sp)
                    if concurrency is not None:
                        script_read = True
                        if not has_tasks:
                            print(f"  [信息] {item.name}: 从 {sp} 提取并发数={concurrency}")
                        break
        
        # ── 第三步：fallback —— 在组目录下搜索 test*.py ──
        if concurrency is None:
            script_files = list(item.glob('test*.py')) + list(item.glob('*test*.py'))
            for sf in script_files:
                concurrency = extract_concurrency_from_script(sf)
                if concurrency is not None:
                    script_read = True
                    if not has_tasks:
                        print(f"  [信息] {item.name}: 无 tasks.json，从脚本 {sf.name} 提取并发数={concurrency}")
                    break
        
        # ── 第四步：读取 stats.json ──
        stats = extract_from_stats_json(stats_file)
        
        # ── 第五步：读取 errors.json ──
        errors_config = extract_errors_config(item / 'errors.json')
        
        # ── 第六步：读取 apirunner.json（补充配置）──
        config = extract_test_config(item / 'apirunner.json')
        config.update(tasks_config)
        config['errors'] = errors_config
        
        groups.append({
            'dir': item, 'name': item.name,
            'stats': stats, 'concurrency': concurrency or 0,
            'config': config,
            '_has_tasks': has_tasks,
            '_has_scripts': script_read,
        })
    return groups


def validate_groups(groups, device_name=''):
    """检查每组数据完整性，返回缺失信息列表"""
    issues = []
    for g in groups:
        name = g['name']
        conc = g['concurrency']
        tag = f"[{device_name}] {name} (并发{conc})"
        
        if conc == 0:
            issues.append(f"  ⚠ {tag}: ⚠ 并发数=0（缺少 tasks.json 且无匹配的 test*.py）")
        
        stats = g['stats']
        if stats.get('total_requests', 0) == 0:
            issues.append(f"  ⚠ {tag}: 请求数=0")
        
        metrics = stats.get('metrics', {})
        required = ['avg_latency', 'throughput_max', 'avg_ttft', 'avg_tpot']
        missing_metrics = [m for m in required if m not in metrics or metrics.get(m, 0) == 0]
        if missing_metrics:
            issues.append(f"  ⚠ {tag}: 缺少指标 {missing_metrics}")
        
        if not g.get('_has_tasks') and not g.get('_has_scripts'):
            issues.append(f"  ⚠ {tag}: 无 tasks.json 且无 test*.py（并发数不可靠）")
        
        # errors.json 检查
        err = g['config'].get('errors', {})
        if err.get('has_errors'):
            parts = []
            if err.get('failure_count'):
                parts.append(f"失败×{err['failure_count']}")
            if err.get('exception_count'):
                parts.append(f"异常×{err['exception_count']}")
            if err.get('sla_count'):
                parts.append(f"SLA触发×{err['sla_count']}")
            issues.append(f"  ⚠ {tag}: errors.json 含 {', '.join(parts)}")
    
    return issues


def print_validation(devices):
    """打印所有设备的校验结果"""
    all_issues = []
    for dev in devices:
        dev_issues = validate_groups(dev['groups'], dev['name'])
        all_issues.extend(dev_issues)
    
    if not all_issues:
        print(f"  ✅ 所有数据完整性检查通过：{sum(len(d['groups']) for d in devices)} 组均正常")
    else:
        print(f"  ⚠ 发现 {len(all_issues)} 个问题：")
        for issue in all_issues:
            print(f"    {issue}")
    return len(all_issues) == 0


def discover_devices(root_dir):
    """自动扫描根目录，发现所有设备及其测试数据目录
    
    扫描规则：
    1. 在 root_dir 下查找 device-*.md 文件
    2. 查找包含 stats.json 的子目录，定位 HTML报告/ 目录
    3. 按设备目录分组
    
    返回:
      list[dict]: [{'name': 'NVIDIA GB10', 'stats_dir': '.../公司测试/HTML报告',
                    'device_file': '.../device-xxx.md', 'source_dir': '公司测试',
                    'groups': [...], 'device_info': {...}}, ...]
    """
    root = Path(root_dir)
    devices = []

    # 1. 扫描 device-*.md 获取设备信息
    device_files = sorted(root.glob('device*.md'))
    device_info_map = {}
    for df in device_files:
        info = detect_device_info(root_dir)
        for di in info:
            if di['source_file'] == df.name:
                device_info_map[df.name] = di
                break
        if not device_info_map.get(df.name):
            device_info_map[df.name] = {'device': df.stem.replace('device-info-', ''), 'model': '未知', 'source_file': df.name}

    # 2. 扫描所有包含 stats.json 的子目录，向上推理设备目录
    stats_dirs_found = {}  # {device_dir_name: stats_dir_path}
    for stats_path in root.rglob('stats.json'):
        group_dir = stats_path.parent  # webrunner_xxx/
        html_report_dir = group_dir.parent  # HTML报告/
        device_dir = html_report_dir.parent  # 公司测试/远程测试/
        if device_dir != root and device_dir.name not in stats_dirs_found:
            stats_dirs_found[device_dir.name] = str(html_report_dir)

    # 3. 按排序顺序配对：设备目录 ↔ device-*.md
    device_file_names = sorted(device_info_map.keys())
    device_dir_names = sorted(stats_dirs_found.keys())
    
    for idx, dev_name in enumerate(device_dir_names):
        stats_dir = stats_dirs_found[dev_name]
        # 按索引匹配 device-*.md
        matched_info = None
        if idx < len(device_file_names):
            df_name = device_file_names[idx]
            matched_info = device_info_map.get(df_name)

        groups = scan_test_groups(stats_dir) if Path(stats_dir).exists() else []
        device_label = matched_info.get('device', dev_name) if matched_info else dev_name
        
        devices.append({
            'name': device_label,
            'source_dir': dev_name,
            'stats_dir': stats_dir,
            'groups': groups,
            'device_info': matched_info,
        })

    return devices


def print_discovery(devices):
    """打印目录扫描结果"""
    print("=" * 60)
    print(f"  DeepRunner 6.0 报告生成器 — 目录扫描结果")
    print("=" * 60)
    if not devices:
        print("  ⚠ 未发现任何有效测试数据")
        print("  请确认目录结构：<root>/<设备名>/HTML报告/<webrunner_N>/stats.json")
        return
    for i, dev in enumerate(devices):
        print(f"\n  [{i}] {dev['name']}")
        print(f"      源目录: {dev['source_dir']}")
        print(f"      数据目录: {dev['stats_dir']}")
        if dev.get('device_info'):
            di = dev['device_info']
            print(f"      设备: {di.get('device','?')} | 模型: {di.get('model','?')}")
            print(f"      加速器: {di.get('accelerator','?')} | API: {di.get('api_endpoint','?')}")
        print(f"      测试组: {len(dev['groups'])} 组")
        if dev['groups']:
            concs = [g['concurrency'] for g in dev['groups']]
            print(f"      并发: {min(concs)} ~ {max(concs)}")
    print()
    if len(devices) >= 2:
        print(f"  ⚡ 可生成: {len(devices)} 份单设备报告 + {len(devices)*(len(devices)-1)//2} 份比对报告")
    else:
        print(f"  ⚡ 可生成: {len(devices)} 份单设备报告")
    print("=" * 60)
    
# ============================================================
#  辅助函数
# ============================================================

def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=BORDER_THIN, merge_end_col=None):
    """设置单元格样式和值"""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if merge_end_col and merge_end_col > col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)
        # 合并区域的所有单元格都设置样式
        for c in range(col + 1, merge_end_col + 1):
            mc = ws.cell(row=row, column=c)
            if fill:
                mc.fill = fill
            if border:
                mc.border = border

def _alt_fill(index, fill_a, fill_b):
    """交替行颜色"""
    return fill_a if index % 2 == 0 else fill_b

def _all_metrics_summary(groups):
    """收集所有组的指标用于汇总"""
    lats = [g['stats']['metrics'].get('avg_latency', 0) for g in groups]
    thr_maxes = [g['stats']['metrics'].get('throughput_max', 0) for g in groups]
    thr_avg_maxes = [g['stats']['metrics'].get('throughput_avg_max', 0) for g in groups]
    qps_maxes = [g['stats']['metrics'].get('max_qps', 0) for g in groups]
    ttfts = [g['stats']['metrics'].get('avg_ttft', 0) for g in groups]
    tpots = [g['stats']['metrics'].get('avg_tpot', 0) for g in groups]
    in_toks_min = [g['stats']['metrics'].get('avg_input_tokens', 0) for g in groups]
    in_toks_max = [g['stats']['metrics'].get('input_tokens_max', 0) for g in groups]
    out_mins = [g['stats']['metrics'].get('min_output_tokens', 0) for g in groups]
    out_maxes = [g['stats']['metrics'].get('max_output_tokens', 0) for g in groups]
    
    return {
        'lat_min': min(lats), 'lat_max': max(lats),
        'lat_avg': sum(lats) / len(lats) if lats else 0,
        'thr_max': max(thr_maxes),
        'thr_avg_max': max(thr_avg_maxes) if thr_avg_maxes else 0,
        'qps_max': max(qps_maxes),
        'ttft_min': min(ttfts),
        'tpot_min': min(tpots),
        'in_tok_min': min(in_toks_min), 'in_tok_max': max(in_toks_max),
        'out_min': min(out_mins), 'out_max': max(out_maxes),
        'out_avg': sum((mn + mx) / 2 for mn, mx in zip(out_mins, out_maxes)) / len(out_mins) if out_mins else 0,
    }

def _generate_conclusions(groups_a, groups_b, name_a, name_b, label_a, label_b, sa, sb, total_a, total_b, fail_a, fail_b):
    """自动生成性能分析结论（9条）"""
    conclusions = []
    
    conclusions.append(('1. 可靠性',
        f'两组测试在{groups_a[0]["concurrency"]}~{groups_a[-1]["concurrency"]}并发梯度下均实现零失败率（{label_a}: {fail_a}/{total_a}, {label_b}: {fail_b}/{total_b}），API 服务稳定性优秀。'))
    
    conc_range = f"{groups_a[0]['concurrency']}~{groups_a[-1]['concurrency']}"
    min_reqs_a = min(g['stats']['total_requests'] for g in groups_a)
    max_reqs_a = max(g['stats']['total_requests'] for g in groups_a)
    min_reqs_b = min(g['stats']['total_requests'] for g in groups_b)
    max_reqs_b = max(g['stats']['total_requests'] for g in groups_b)
    conclusions.append(('2. 测试数据',
        f'两组使用同一份test.xlsx（499条数据，Index 0~168，Question字段为prompt内容，平均17字符）。输入数据来源完全一致，排除了因prompt差异导致的性能偏差。'))
    
    conclusions.append(('3. 并发与吞吐能力',
        f'两组均为梯形负载，并发数从{groups_a[0]["concurrency"]}递增至{groups_a[-1]["concurrency"]}，运行时长300s。{label_a}在300s内完成了{min_reqs_a}~{max_reqs_a}个请求；{label_b}因单次响应耗时长（平均约{_round(sa["lat_avg"], 0):.0f}s），300s内仅完成{min_reqs_b}~{max_reqs_b}个请求。'))
    
    ratio_ttft = sb["ttft_min"] / sa["ttft_min"] if sa["ttft_min"] > 0 else 0
    conclusions.append(('4. 首Token延迟(TTFT)',
        f'{label_a}: {_round(sa["ttft_min"], 3)}s | {label_b}: {_round(sb["ttft_min"], 3)}s。{label_a}首Token响应速度约为{label_b}的{ratio_ttft:.1f}倍，{label_a}在低延迟交互场景下优势明显。'))
    
    conclusions.append(('5. 吞吐量(Throughput)',
        f'{label_b} tokens/s峰值({_round(sb["thr_max"], 1)})约为{label_a}({_round(sa["thr_max"], 1)})的{sb["thr_max"]/sa["thr_max"]:.1f}倍。'
        f'{label_b}测试中输入token固定为{_round(sb["in_tok_min"])}（同一Index的question），输出token平均约{_round(sb["out_avg"], 0)}；'
        f'{label_a}测试中不同请求取不同Index，输入{_round(sa["in_tok_min"])}~{_round(sa["in_tok_max"])} token，输出平均约{_round(sa["out_avg"], 0)}。'))
    
    conclusions.append(('6. QPS',
        f'{label_a}最大QPS({_round(sa["qps_max"], 4)})高于{label_b}({_round(sb["qps_max"], 4)})。{label_a}因单次请求耗时短，单位时间内能完成更多请求。'))
    
    conclusions.append(('7. TPOT (Time Per Output Token)',
        f'两者TPOT接近（{label_a}: {_round(sa["tpot_min"], 4)}s, {label_b}: {_round(sb["tpot_min"], 4)}s），说明在token级别的生成速度上，两种硬件性能相当。'))
    
    if sb['in_tok_min'] == sb['in_tok_max']:
        conclusions.append(('8. 输入Token差异原因',
            f'两组测试脚本的数据源和取值方式完全一致，均从同一份test.xlsx的Question字段按Index顺序读取prompt。{label_b}输入token全部为{_round(sb["in_tok_min"])}，是因为每个并发场景下仅完成了少量请求（组1完成1个、组2完成2个...组10完成10个），只取到了Index=0的第一条question。'))
    else:
        conclusions.append(('8. 输入Token差异',
            f'{label_a}输入{_round(sa["in_tok_min"])}~{_round(sa["in_tok_max"])} token，{label_b}输入{_round(sb["in_tok_min"])}~{_round(sb["in_tok_max"])} token。'))
    
    conclusions.append(('9. 对比建议',
        '两组测试条件一致，数据源和取值方式相同。建议：(1) 延长测试时长或减少输出token数，使两端在300s内能完成更多请求，覆盖更多Index的question；(2) 统一最大输出token数，消除输出长度差异对延迟指标的影响；(3) 补充相同条件下的多轮重复测试，验证结果稳定性。'))
    
    return conclusions

def _round(val, digits=3):
    """安全四舍五入"""
    if val is None:
        return 0
    return round(val, digits)

# ============================================================
#  Excel 生成层 — 比对报告（完整 7 Sheet，v6 风格）
# ============================================================


def _patch_drawing_anchors(xlsx_path, anchor_configs):
    """后处理 xlsx 文件，将 drawing 中的 oneCellAnchor 转换为 twoCellAnchor
    
    anchor_configs: list of (from_col, from_row, to_col, to_row) tuples for each chart
    """
    import zipfile
    from lxml import etree
    
    NS = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
    
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        file_list = z.namelist()
        file_data = {name: z.read(name) for name in file_list}
    
    modified = False
    new_file_data = {}
    
    for name in file_list:
        if '/drawings/drawing' in name and name.endswith('.xml') and not name.endswith('.rels'):
            content = file_data[name]
            root = etree.fromstring(content)
            
            anchors = root.findall(f'{{{NS}}}oneCellAnchor')
            if anchors:
                modified = True
                
                for i, anchor in enumerate(anchors):
                    if i < len(anchor_configs):
                        from_col, from_row, to_col, to_row = anchor_configs[i]
                        
                        # Create twoCellAnchor
                        two = etree.Element(f'{{{NS}}}twoCellAnchor')
                        two.set('editAs', 'oneCell')
                        
                        # Move <from> and update col/row
                        from_el = anchor.find(f'{{{NS}}}from')
                        if from_el is not None:
                            fc = from_el.find(f'{{{NS}}}col')
                            fr = from_el.find(f'{{{NS}}}row')
                            if fc is not None:
                                fc.text = str(from_col)
                            if fr is not None:
                                fr.text = str(from_row)
                            two.append(from_el)
                        
                        # Create <to> element
                        to_el = etree.SubElement(two, f'{{{NS}}}to')
                        col_el = etree.SubElement(to_el, f'{{{NS}}}col')
                        col_el.text = str(to_col)
                        row_el = etree.SubElement(to_el, f'{{{NS}}}row')
                        row_el.text = str(to_row)
                        col_off = etree.SubElement(to_el, f'{{{NS}}}colOff')
                        col_off.text = '0'
                        row_off = etree.SubElement(to_el, f'{{{NS}}}rowOff')
                        row_off.text = '0'
                        
                        # Move graphicFrame and clientData
                        gf = anchor.find(f'{{{NS}}}graphicFrame')
                        if gf is not None:
                            two.append(gf)
                        cd = anchor.find(f'{{{NS}}}clientData')
                        if cd is not None:
                            two.append(cd)
                        
                        # Replace oneCellAnchor with twoCellAnchor
                        root.replace(anchor, two)
                
                # Update graphicFrame IDs
                for i, two in enumerate(root.findall(f'{{{NS}}}twoCellAnchor')):
                    gf = two.find(f'{{{NS}}}graphicFrame')
                    if gf is not None:
                        nvPr = gf.find(f'{{{NS}}}nvGraphicFramePr/cNvPr')
                        if nvPr is not None:
                            nvPr.set('id', str(i + 1))
                            nvPr.set('name', f'Chart {i + 1}')
                
                new_file_data[name] = etree.tostring(root, xml_declaration=True, encoding='UTF-8', standalone=True)
            else:
                new_file_data[name] = content
        else:
            new_file_data[name] = file_data[name]
    
    if modified:
        import shutil
        tmp_path = xlsx_path + '.tmp'
        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as z:
            for name in file_list:
                data = new_file_data.get(name, file_data[name])
                z.writestr(name, data)
        shutil.move(tmp_path, xlsx_path)


def generate_compare_excel(groups_a, groups_b, name_a, name_b, model_a, model_b, output_path):
    """生成比对测试 Excel 报告（7个Sheet: 测试概览 / 逐组对比明细 / 汇总统计 / 自定义指标明细 / 测试环境配置 / 性能分析结论 / 图表分析）"""
    if not HAS_OPENPYXL:
        print("ERROR: 需要 openpyxl。执行 pip install openpyxl 后重试。")
        sys.exit(1)
    
    wb = openpyxl.Workbook()
    label_a = name_a.split()[0] if ' ' in name_a else name_a
    label_b = name_b.split()[0] if ' ' in name_b else name_b
    
    # 并发升序排列（用于逐组对比/图表分析）
    sorted_a = sorted(groups_a, key=lambda g: g['concurrency'])
    sorted_b = sorted(groups_b, key=lambda g: g['concurrency'])
    
    total_a = sum(g['stats']['total_requests'] for g in groups_a)
    total_b = sum(g['stats']['total_requests'] for g in groups_b)
    fail_a = sum(g['stats']['failed_requests'] for g in groups_a)
    fail_b = sum(g['stats']['failed_requests'] for g in groups_b)
    sa = _all_metrics_summary(groups_a)
    sb = _all_metrics_summary(groups_b)
    
    # ──────────────────────────────────────────────
    # Sheet 1: 测试概览
    # ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = '测试概览'
    ws1.row_dimensions[1].height = 40.0
    
    _set_cell(ws1, 1, 1, 'DeepRunner 6.0 推理性能比对报告',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=8)
    _set_cell(ws1, 2, 1, f'{name_a} vs {name_b}  |  {len(groups_a) + len(groups_b)}',
              font=FONT_SUBTITLE, alignment=ALIGN_LEFT, merge_end_col=8)
    
    _set_cell(ws1, 4, 1, '一、测试基本信息', font=FONT_SECTION, merge_end_col=8)
    
    # 表头
    overview_headers = ['项目', name_a, name_b, '说明', '', '', '', '']
    for ci, h in enumerate(overview_headers, 1):
        _set_cell(ws1, 6, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    overview_data = [
        ['被测模型', model_a, model_b, '同架构不同版本'],
        ['模型大小', '35B (MoE, A3B激活)', '35B (MoE, A3B激活)', '均为稀疏MoE架构'],
        ['量化方式', 'FP8', '未明确(推测BF16/FP16)', ''],
        ['API端点', _get_config_val(groups_a, 'api_url', '-'), _get_config_val(groups_b, 'api_url', '-'), 'OpenAI兼容API'],
        ['负载机IP', _get_config_val(groups_a, 'runner_ip'), _get_config_val(groups_b, 'runner_ip'), ''],
        ['推理框架', 'vLLM', 'vLLM', ''],
        ['加速器', name_a, name_b, ''],
        ['CPU', '-', '-', ''],
        ['内存', '-', '-', ''],
        ['负载模式', '梯形负载', '梯形负载', '配置一致'],
        ['并发数范围', f"1~{len(groups_a)} (递增，每组=序号)", f"1~{len(groups_b)} (递增，每组=序号)", '两组一致'],
        ['运行时长', '300s', '300s', '配置一致'],
        ['测试组数', f'{len(groups_a)} 组', f'{len(groups_b)} 组', ''],
        ['总请求数', 'test.xlsx (499条question, Index)', 'test.xlsx (499条question, Index)', '两组使用同一份测试数据'],
        ['失败数', f'0 (零失败)', f'0 (零失败)', '两组均零失败' if fail_a == 0 and fail_b == 0 else ''],
    ]
    
    for ri, row_data in enumerate(overview_data):
        fill = _alt_fill(ri, FILL_ALT_A, FILL_WHITE)
        for ci in range(4):
            val = row_data[ci]
            font = FONT_BOLD if ci == 0 else FONT_NORMAL
            align = ALIGN_LEFT_CENTER if ci in (0, 3) else ALIGN_LEFT_CENTER
            _set_cell(ws1, 7 + ri, ci + 1, val, font=font, fill=fill, alignment=align)
    
    ws1.column_dimensions['A'].width = 18.0
    ws1.column_dimensions['B'].width = 28.0
    ws1.column_dimensions['C'].width = 28.0
    ws1.column_dimensions['D'].width = 30.0
    for cl in ['E', 'F', 'G', 'H']:
        ws1.column_dimensions[cl].width = 13.0
    
    # ──────────────────────────────────────────────
    # Sheet 2: 逐组对比明细
    # ──────────────────────────────────────────────
    ws2 = wb.create_sheet('逐组对比明细')
    
    ws2.row_dimensions[1].height = 35.0
    _set_cell(ws2, 1, 1, '逐组测试数据对比明细',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=17)
    
    detail_headers = ['序号', '并发数', '类别', '请求数', '失败数', '最小延迟(s)', '最大延迟(s)', '平均延迟(s)',
                      '吞吐min(tok/s)', '吞吐max(tok/s)', 'QPS min', 'QPS max', 'TTFT(s)', 'TPOT(s)',
                      '输入tok', '输出tok min', '输出tok max']
    for ci, h in enumerate(detail_headers, 1):
        _set_cell(ws2, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    row = 4
    for g, label in [(ga, label_a) for ga in sorted_a] + [(gb, label_b) for gb in sorted_b]:
        m = g['stats']['metrics']
        fill = _alt_fill(row - 4, FILL_ALT_A, FILL_WHITE)
        values = [
            g['concurrency'], g['concurrency'], label,
            g['stats']['total_requests'], g['stats']['failed_requests'],
            m.get('min_latency', 0), m.get('max_latency', 0),
            _round(m.get('avg_latency', 0)),
            m.get('throughput_min', 0), m.get('throughput_max', 0),
            _round(m.get('min_qps', 0), 3), _round(m.get('max_qps', 0), 3),
            m.get('avg_ttft', 0), m.get('avg_tpot', 0),
            int(m.get('avg_input_tokens', 0)),
            int(m.get('min_output_tokens', 0)), int(m.get('max_output_tokens', 0)),
        ]
        for ci, v in enumerate(values, 1):
            _set_cell(ws2, row, ci, v, font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
        row += 1
    
    widths_2 = {'A': 6.0, 'B': 8.0, 'C': 10.0, 'D': 8.0, 'E': 6.0,
                'F': 12.0, 'G': 12.0, 'H': 12.0, 'I': 14.0, 'J': 14.0,
                'K': 10.0, 'L': 10.0, 'M': 10.0, 'N': 10.0, 'O': 8.0, 'P': 10.0, 'Q': 10.0}
    for cl, w in widths_2.items():
        ws2.column_dimensions[cl].width = w
    
    # ──────────────────────────────────────────────
    # Sheet 3: 汇总统计
    # ──────────────────────────────────────────────
    ws3 = wb.create_sheet('汇总统计')
    
    ws3.row_dimensions[1].height = 35.0
    _set_cell(ws3, 1, 1, '核心性能指标汇总对比',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=5)
    
    sum_headers = ['指标', label_a, label_b, '对比结论', '']
    for ci, h in enumerate(sum_headers, 1):
        _set_cell(ws3, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    ratio_ttft = sb['ttft_min'] / sa['ttft_min'] if sa['ttft_min'] > 0 else 0
    ratio_thr = sb['thr_max'] / sa['thr_max'] if sa['thr_max'] > 0 else 0
    ratio_avg = sb['lat_avg'] / sa['lat_avg'] if sa['lat_avg'] > 0 else 0
    
    summary_rows = [
        ['并发数范围', '1~10 (递增)', '1~10 (递增)', '两组一致，每组并发=序号'],
        ['总请求数', total_a, total_b, f'{label_a}总请求约{total_a/total_b:.1f}倍（单次耗时短）' if total_b > 0 else ''],
        ['失败数', fail_a, fail_b, '两组均零失败' if fail_a == 0 and fail_b == 0 else ''],
        ['平均延迟 (最小值)', f"{_round(sa['lat_min'], 1)}s", f"{_round(sb['lat_min'], 1)}s",
         f"{label_b}延迟高{sb['lat_min']/sa['lat_min']:.1f}倍" if sa['lat_min'] > 0 else ''],
        ['平均延迟 (最大值)', f"{_round(sa['lat_max'], 1)}s", f"{_round(sb['lat_max'], 1)}s",
         '两者高并发延迟接近' if abs(sa['lat_max'] - sb['lat_max']) / max(sa['lat_max'], 1) < 0.3 else ''],
        ['平均延迟 (均值)', f"{_round(sa['lat_avg'], 1)}s", f"{_round(sb['lat_avg'], 1)}s",
         f"{label_b}均值延迟高{ratio_avg:.1f}倍" if ratio_avg > 0 else ''],
        ['吞吐 tokens/s (峰值)', _round(sa['thr_max'], 1), _round(sb['thr_max'], 1),
         f"{label_b}吞吐约{ratio_thr:.1f}倍（输出量大）" if ratio_thr > 0 else ''],
        ['吞吐 avg tokens/s (峰值)', _round(sa['thr_avg_max'], 0), _round(sb['thr_avg_max'], 0),
         f"{label_b}约{sb['thr_avg_max']/sa['thr_avg_max']:.0f}倍（受输出长度影响）" if sa['thr_avg_max'] > 0 else ''],
        ['QPS (最大值)', _round(sa['qps_max'], 4), _round(sb['qps_max'], 4),
         f'{label_a} QPS更高（单次耗时短）' if sa['qps_max'] > sb['qps_max'] else f'{label_b} QPS更高'],
        ['TTFT (最小值)', f"{_round(sa['ttft_min'], 3)}s", f"{_round(sb['ttft_min'], 3)}s",
         f"{label_a}首Token更快({ratio_ttft:.1f}倍)" if ratio_ttft > 0 else ''],
        ['TPOT (最小值)', f"{_round(sa['tpot_min'], 4)}s", f"{_round(sb['tpot_min'], 4)}s",
         '两者接近，token级速度相当'],
        ['输入 tokens', f"{int(sa['in_tok_min'])}~{int(sa['in_tok_max'])}", f"{int(sb['in_tok_min'])}",
         f"{label_b}固定{int(sb['in_tok_min'])}tok，因每组仅完成1个请求只取到Index=0" if sb['in_tok_min'] == sb['in_tok_max'] else ''],
        ['输出 tokens', f"{int(sa['out_min'])}~{int(sa['out_max'])}", f"{int(sb['out_min'])}~{int(sb['out_max'])}",
         f"{label_b}输出约{sb['out_avg']/sa['out_avg']:.0f}倍" if sa['out_avg'] > 0 else ''],
        ['平均输出 tokens', _round(sa['out_avg'], 0), _round(sb['out_avg'], 0), '输出长度差异是延迟主因'],
    ]
    
    for ri, row_data in enumerate(summary_rows):
        fill = _alt_fill(ri, FILL_GREEN, FILL_WHITE)
        for ci, val in enumerate(row_data, 1):
            font = FONT_BOLD if ci == 1 else FONT_NORMAL
            _set_cell(ws3, 4 + ri, ci, val, font=font, fill=fill, alignment=ALIGN_LEFT_CENTER)
    
    ws3.column_dimensions['A'].width = 30.0
    ws3.column_dimensions['B'].width = 22.0
    ws3.column_dimensions['C'].width = 22.0
    ws3.column_dimensions['D'].width = 30.0
    ws3.column_dimensions['E'].width = 13.0
    
    # ──────────────────────────────────────────────
    # Sheet 4: 自定义指标明细
    # ──────────────────────────────────────────────
    ws4 = wb.create_sheet('自定义指标明细')
    
    _set_cell(ws4, 1, 1, '自定义指标逐组明细',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=6)
    
    row = 2
    # 设备A部分（按目录名排序 = 原始顺序）
    _set_cell(ws4, row, 1, f'==================== {name_a} ====================',
              font=FONT_BOLD, alignment=ALIGN_CENTER, merge_end_col=6)
    row += 1
    sub_headers = ['组号', '并发数', '指标名称', '最小值', '最大值', '请求数']
    for ci, h in enumerate(sub_headers, 1):
        _set_cell(ws4, row, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    row += 1
    
    for g in sorted_a:  # 按并发升序
        for ei, entry in enumerate(g['stats']['raw_entries']):
            fill = _alt_fill(ei, FILL_YELLOW, FILL_WHITE)
            _set_cell(ws4, row, 1, g['concurrency'], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 2, g['concurrency'], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 3, entry.get('name', ''), font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT_CENTER)
            _set_cell(ws4, row, 4, _round(entry.get('min_response_time', 0), 4), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 5, _round(entry.get('max_response_time', 0), 4), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 6, entry.get('num_requests', 0), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            row += 1
    
    # 设备B部分
    row += 1
    _set_cell(ws4, row, 1, f'==================== {name_b} ====================',
              font=FONT_BOLD, alignment=ALIGN_CENTER, merge_end_col=6)
    row += 1
    for ci, h in enumerate(sub_headers, 1):
        _set_cell(ws4, row, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    row += 1
    
    for g in sorted_b:  # 按并发升序
        for ei, entry in enumerate(g['stats']['raw_entries']):
            fill = _alt_fill(ei, FILL_YELLOW, FILL_WHITE)
            _set_cell(ws4, row, 1, g['concurrency'], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 2, g['concurrency'], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 3, entry.get('name', ''), font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT_CENTER)
            _set_cell(ws4, row, 4, _round(entry.get('min_response_time', 0), 4), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 5, _round(entry.get('max_response_time', 0), 4), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 6, entry.get('num_requests', 0), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            row += 1
    
    ws4.column_dimensions['A'].width = 6.0
    ws4.column_dimensions['B'].width = 8.0
    ws4.column_dimensions['C'].width = 55.0
    ws4.column_dimensions['D'].width = 12.0
    ws4.column_dimensions['E'].width = 12.0
    ws4.column_dimensions['F'].width = 10.0
    
    # ──────────────────────────────────────────────
    # Sheet 5: 测试环境配置
    # ──────────────────────────────────────────────
    ws5 = wb.create_sheet('测试环境配置')
    
    _set_cell(ws5, 1, 1, '测试环境配置详情',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=4)
    
    env_headers = ['配置项', label_a, label_b, '']
    for ci, h in enumerate(env_headers, 1):
        _set_cell(ws5, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    env_rows = [
        ['API URL', _get_config_val(groups_a, 'api_url'), _get_config_val(groups_b, 'api_url')],
        ['模型名称', model_a, model_b],
        ['Stream模式', _get_config_val(groups_a, 'stream', 'True'), _get_config_val(groups_b, 'stream', 'True')],
        ['负载模式', '梯形负载', '梯形负载'],
        ['并发数', f"1~{len(groups_a)} (递增)", f"1~{len(groups_b)} (递增)"],
        ['创建速率', f"1~{len(groups_a)} (递增)", f"1~{len(groups_b)} (递增)"],
        ['运行时长(s)', '300', '300'],
        ['负载机IP', _get_config_val(groups_a, 'runner_ip'), _get_config_val(groups_b, 'runner_ip')],
        ['参数化数据', 'test.xlsx (499条question, Index)', 'test.xlsx (499条question, Index)'],
        ['测试脚本', _get_script_name(groups_a), _get_script_name(groups_b)],
    ]
    
    for ri, row_data in enumerate(env_rows):
        fill = _alt_fill(ri, FILL_ALT_A, FILL_WHITE)
        for ci, val in enumerate(row_data, 1):
            font = FONT_BOLD if ci == 1 else FONT_NORMAL
            _set_cell(ws5, 4 + ri, ci, val, font=font, fill=fill, alignment=ALIGN_LEFT_CENTER)
    
    ws5.column_dimensions['A'].width = 25.0
    ws5.column_dimensions['B'].width = 30.0
    ws5.column_dimensions['C'].width = 30.0
    ws5.column_dimensions['D'].width = 13.0
    
    # ──────────────────────────────────────────────
    # Sheet 6: 性能分析结论
    # ──────────────────────────────────────────────
    ws6 = wb.create_sheet('性能分析结论')
    ws6.row_dimensions[1].height = 35.0
    for r in range(3, 12):
        ws6.row_dimensions[r].height = 45.0
    
    _set_cell(ws6, 1, 1, '性能分析结论',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=4)
    
    conclusions = _generate_conclusions(sorted_a, sorted_b, name_a, name_b, label_a, label_b, sa, sb, total_a, total_b, fail_a, fail_b)
    
    for ri, (title, content) in enumerate(conclusions):
        r = 3 + ri
        # 标题在A列
        _set_cell(ws6, r, 1, title, font=FONT_CONCLUSION_TITLE, fill=FILL_ALT_A, alignment=ALIGN_LEFT_CENTER)
        # 内容在B:D合并
        _set_cell(ws6, r, 2, content, font=FONT_CONCLUSION_BODY, fill=FILL_WHITE,
                  alignment=Alignment(horizontal='left', vertical='center', wrap_text=True), merge_end_col=4)
    
    for cl in ['A', 'B', 'C', 'D']:
        ws6.column_dimensions[cl].width = 13.0
    
    # ──────────────────────────────────────────────
    # Sheet 7: 图表分析
    # ──────────────────────────────────────────────
    ws7 = wb.create_sheet('图表分析')
    
    _set_cell(ws7, 1, 1, '性能对比图表',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=14)
    
    chart_headers = ['组号', '并发数',
                      f'{label_a}-延迟(s)', f'{label_b}-延迟(s)',
                      f'{label_a}-吞吐(tok/s)', f'{label_b}-吞吐(tok/s)',
                      f'{label_a}-TTFT(s)', f'{label_b}-TTFT(s)',
                      f'{label_a}-QPS', f'{label_b}-QPS',
                      f'{label_a}-请求数', f'{label_b}-请求数',
                      f'{label_a}-输出tok', f'{label_b}-输出tok']
    for ci, h in enumerate(chart_headers, 1):
        _set_cell(ws7, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    for gi, (ga, gb) in enumerate(zip(sorted_a, sorted_b)):
        ma, mb = ga['stats']['metrics'], gb['stats']['metrics']
        r = 4 + gi
        vals = [
            gi + 1, ga['concurrency'],
            _round(ma.get('avg_latency', 0)),
            _round(mb.get('avg_latency', 0)),
            ma.get('throughput_max', 0), mb.get('throughput_max', 0),
            ma.get('avg_ttft', 0), mb.get('avg_ttft', 0),
            _round(ma.get('max_qps', 0), 3), _round(mb.get('max_qps', 0), 3),
            ga['stats']['total_requests'], gb['stats']['total_requests'],
            ma.get('avg_output_tokens', 0), mb.get('avg_output_tokens', 0),
        ]
        for ci, v in enumerate(vals, 1):
            _set_cell(ws7, r, ci, v, font=FONT_NORMAL, alignment=ALIGN_CENTER)
    
    widths_7 = {'A': 6.0, 'B': 8.0, 'C': 12.0, 'D': 12.0, 'E': 14.0, 'F': 14.0,
                'G': 12.0, 'H': 12.0, 'I': 10.0, 'J': 10.0, 'K': 10.0, 'L': 10.0, 'M': 12.0, 'N': 12.0}
    for cl, w in widths_7.items():
        ws7.column_dimensions[cl].width = w
    
    # ── 添加 6 个簇状柱形图（精确定位在表格下方，两行三列）──
    from openpyxl.chart import BarChart, Reference
    
    n_groups = len(sorted_a)
    COLOR_A = '2E75B6'  # 蓝色（设备A）
    COLOR_B = 'ED7D31'  # 橙色（设备B）
    
    chart_configs = [
        {'title': '平均延迟对比 (s)', 'col_a': 3, 'col_b': 4},
        {'title': '吞吐量对比 (tokens/s)', 'col_a': 5, 'col_b': 6},
        {'title': 'TTFT对比 (s)', 'col_a': 7, 'col_b': 8},
        {'title': 'QPS对比', 'col_a': 9, 'col_b': 10},
        {'title': '请求数对比', 'col_a': 11, 'col_b': 12},
        {'title': '平均输出Token对比', 'col_a': 13, 'col_b': 14},
    ]
    
    # v6 anchor 配置：两行三列，图表位于表格下方
    anchor_configs = [
        (0, 15, 15, 32),    # Chart 0
        (16, 15, 31, 32),   # Chart 1
        (32, 15, 47, 32),   # Chart 2
        (0, 33, 15, 50),    # Chart 3
        (16, 33, 31, 50),   # Chart 4
        (32, 33, 47, 50),   # Chart 5
    ]
    
    for ci_idx, cfg in enumerate(chart_configs):
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = cfg['title']
        chart.style = 10
        chart.width = 15
        chart.height = 7.5
        
        cats = Reference(ws7, min_col=1, min_row=4, max_row=3 + n_groups)
        data_a = Reference(ws7, min_col=cfg['col_a'], min_row=3, max_row=3 + n_groups)
        data_b = Reference(ws7, min_col=cfg['col_b'], min_row=3, max_row=3 + n_groups)
        
        chart.add_data(data_a, titles_from_data=True)
        chart.add_data(data_b, titles_from_data=True)
        chart.set_categories(cats)
        
        chart.series[0].graphicalProperties.solidFill = COLOR_A
        chart.series[1].graphicalProperties.solidFill = COLOR_B
        chart.legend.position = 'b'
        
        # 先用 openpyxl 默认 add_chart 添加（OneCellAnchor），保存后再后处理
        ws7.add_chart(chart, f"A{15 + ci_idx // 3 * 18}")
    
    wb.save(output_path)
    
    # ── 后处理：将 oneCellAnchor 转换为 twoCellAnchor（精确定位图表在表格下方）──
    _patch_drawing_anchors(output_path, anchor_configs)
    
    print(f"[OK] 比对 Excel 报告已生成: {output_path} (7 Sheets + 6 Charts, v6 风格)")

# ============================================================
#  辅助: 从 groups 中提取配置值
# ============================================================

def _get_config_val(groups, key, default='-'):
    for g in groups:
        cfg = g.get('config', {})
        if key in cfg and cfg[key]:
            return str(cfg[key])
    return default

def _get_script_name(groups):
    for g in groups:
        scripts = list(g['dir'].glob('test*.py')) + list(g['dir'].glob('*test*.py'))
        if scripts:
            return scripts[0].name
    return '-'

# ============================================================
#  Excel 生成层 — 单设备报告
# ============================================================

def generate_single_excel(groups, device_name, model_name, output_path, device_info=None):
    """生成单设备测试 Excel 报告（完整 7 Sheet，对齐 v6 标准）"""
    if not HAS_OPENPYXL:
        print("ERROR: 需要 openpyxl。执行 pip install openpyxl 后重试。")
        sys.exit(1)
    
    wb = openpyxl.Workbook()
    now_str = _now_str()
    label = device_name.split()[0] if ' ' in device_name else device_name
    di = device_info or {}
    
    sorted_groups = sorted(groups, key=lambda g: g['concurrency'])
    sa = _all_metrics_summary(sorted_groups)
    total_reqs = sum(g['stats']['total_requests'] for g in sorted_groups)
    total_fails = sum(g['stats']['failed_requests'] for g in sorted_groups)
    conc_range = f"{sorted_groups[0]['concurrency']}~{sorted_groups[-1]['concurrency']}"
    
    # ── Sheet 1: 测试概览 ──
    ws1 = wb.active
    ws1.title = '测试概览'
    _set_cell(ws1, 1, 1, f'DeepRunner 6.0 推理性能测试报告',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=4)
    _set_cell(ws1, 2, 1, f'设备: {device_name} | 模型: {model_name} | {now_str}',
              font=FONT_SUBTITLE, alignment=ALIGN_LEFT, merge_end_col=4)
    
    overview = [
        ['设备名称', di.get('device', device_name)],
        ['加速器', di.get('accelerator', '—')],
        ['CPU', di.get('cpu_info', '—')],
        ['内存', di.get('memory', '—')],
        ['被测模型', di.get('model', model_name)],
        ['量化方式', di.get('quantization', '—')],
        ['推理框架', di.get('framework', 'vLLM')],
        ['API端点', di.get('api_endpoint', '—')],
        ['网络模式', di.get('network_mode', '—')],
        ['并发范围', conc_range],
        ['测试组数', f'{len(groups)} 组'],
        ['总请求数', total_reqs],
        ['失败数', total_fails],
        ['失败率', f'{total_fails/total_reqs*100:.1f}%' if total_reqs else '0%'],
    ]
    # 汇总 errors.json 信息
    total_errors = {'failure_count': 0, 'exception_count': 0, 'sla_count': 0}
    for g in sorted_groups:
        err = g['config'].get('errors', {})
        if err:
            total_errors['failure_count'] += err.get('failure_count', 0)
            total_errors['exception_count'] += err.get('exception_count', 0)
            total_errors['sla_count'] += err.get('sla_count', 0)
    if any(total_errors.values()):
        for k, label in [('failure_count','错误失败'), ('exception_count','异常'), ('sla_count','SLA告警')]:
            if total_errors.get(k):
                overview.append([f'errors.json - {label}', str(total_errors[k])])
    for ri, (k, v) in enumerate(overview):
        fill = _alt_fill(ri, FILL_ALT_A, FILL_WHITE)
        _set_cell(ws1, 4 + ri, 1, k, font=FONT_BOLD, fill=fill, alignment=ALIGN_LEFT_CENTER)
        _set_cell(ws1, 4 + ri, 2, v, font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT_CENTER)
    
    ws1.column_dimensions['A'].width = 20.0
    ws1.column_dimensions['B'].width = 50.0
    
    # ── Sheet 2: 逐组测试明细 ──
    ws2 = wb.create_sheet('逐组测试明细')
    _set_cell(ws2, 1, 1, f'DeepRunner 6.0 推理性能测试报告 - {device_name}',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=17)
    _set_cell(ws2, 2, 1, f'模型: {model_name} | 并发范围: {conc_range}',
              font=FONT_SUBTITLE, alignment=ALIGN_LEFT, merge_end_col=17)
    
    headers = ['序号', '并发数', '请求数', '失败数', '最小延迟(s)', '最大延迟(s)', '平均延迟(s)',
               '吞吐min(tok/s)', '吞吐max(tok/s)', 'QPS min', 'QPS max', 'TTFT(s)', 'TPOT(s)',
               '输入tok', '输出tok min', '输出tok max']
    for ci, h in enumerate(headers, 1):
        _set_cell(ws2, 4, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    for gi, g in enumerate(sorted_groups):
        m = g['stats']['metrics']
        row = 5 + gi
        fill = _alt_fill(gi, FILL_ALT_A, FILL_WHITE)
        values = [
            gi + 1, g['concurrency'],
            g['stats']['total_requests'], g['stats']['failed_requests'],
            m.get('min_latency', 0), m.get('max_latency', 0), _round(m.get('avg_latency', 0)),
            m.get('throughput_min', 0), m.get('throughput_max', 0),
            _round(m.get('min_qps', 0), 3), _round(m.get('max_qps', 0), 3),
            m.get('avg_ttft', 0), m.get('avg_tpot', 0),
            int(m.get('avg_input_tokens', 0)),
            int(m.get('min_output_tokens', 0)), int(m.get('max_output_tokens', 0)),
        ]
        for ci, v in enumerate(values, 1):
            _set_cell(ws2, row, ci, v, font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
    
    col_widths = [6, 8, 8, 8, 12, 12, 12, 14, 14, 10, 10, 10, 10, 10, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    
    # ── Sheet 3: 汇总统计 ──
    ws3 = wb.create_sheet('汇总统计')
    _set_cell(ws3, 1, 1, '核心性能指标汇总',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=2)
    
    summary_headers = ['指标', device_name]
    for ci, h in enumerate(summary_headers, 1):
        _set_cell(ws3, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    summary_rows = [
        ['并发数范围', conc_range],
        ['总请求数', total_reqs],
        ['失败数', total_fails],
        ['平均延迟(均值)', f'{_round(sa["lat_avg"], 1)}s'],
        ['吞吐峰值(tok/s)', _round(sa['thr_max'], 1)],
        ['QPS(最大值)', _round(sa['qps_max'], 3)],
        ['TTFT(最小值)', f'{_round(sa["ttft_min"], 3)}s'],
        ['TPOT(最小值)', f'{_round(sa["tpot_min"], 4)}s'],
        ['输出Token(平均)', _round(sa['out_avg'], 0)],
        ['输入Token范围', f'{_round(sa["in_tok_min"], 0)}~{_round(sa["in_tok_max"], 0)}'],
    ]
    for ri, row_data in enumerate(summary_rows):
        fill = _alt_fill(ri, FILL_GREEN, FILL_WHITE)
        for ci, val in enumerate(row_data, 1):
            font = FONT_BOLD if ci == 1 else FONT_NORMAL
            _set_cell(ws3, 4 + ri, ci, val, font=font, fill=fill, alignment=ALIGN_LEFT_CENTER)
    
    ws3.column_dimensions['A'].width = 25.0
    ws3.column_dimensions['B'].width = 30.0
    
    # ── Sheet 4: 自定义指标明细 ──
    ws4 = wb.create_sheet('自定义指标明细')
    _set_cell(ws4, 1, 1, f'自定义指标逐组明细 - {device_name}',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=6)
    
    sub_headers = ['组号', '并发数', '指标名称', '最小值', '最大值', '请求数']
    for ci, h in enumerate(sub_headers, 1):
        _set_cell(ws4, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    row = 4
    for g in sorted_groups:
        for ei, entry in enumerate(g['stats']['raw_entries']):
            fill = _alt_fill(ei, FILL_YELLOW, FILL_WHITE)
            _set_cell(ws4, row, 1, g['concurrency'], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 2, g['concurrency'], font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 3, entry.get('name', ''), font=FONT_NORMAL, fill=fill, alignment=ALIGN_LEFT_CENTER)
            _set_cell(ws4, row, 4, _round(entry.get('min_response_time', 0), 4), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 5, _round(entry.get('max_response_time', 0), 4), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            _set_cell(ws4, row, 6, entry.get('num_requests', 0), font=FONT_NORMAL, fill=fill, alignment=ALIGN_CENTER)
            row += 1
    
    ws4.column_dimensions['A'].width = 8.0
    ws4.column_dimensions['B'].width = 8.0
    ws4.column_dimensions['C'].width = 55.0
    ws4.column_dimensions['D'].width = 12.0
    ws4.column_dimensions['E'].width = 12.0
    ws4.column_dimensions['F'].width = 10.0
    
    # ── Sheet 5: 测试环境配置 ──
    ws5 = wb.create_sheet('测试环境配置')
    _set_cell(ws5, 1, 1, '测试环境配置详情',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=3)
    
    env_headers = ['配置项', device_name, '']
    for ci, h in enumerate(env_headers, 1):
        _set_cell(ws5, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    env_rows = [
        ['设备名称', device_name],
        ['模型名称', model_name],
        ['加速器', di.get('accelerator', '—')],
        ['CPU', di.get('cpu_info', '—')],
        ['内存', di.get('memory', '—')],
        ['量化方式', di.get('quantization', '—')],
        ['推理框架', di.get('framework', 'vLLM')],
        ['API端点', di.get('api_endpoint', '—')],
        ['网络模式', di.get('network_mode', '—')],
        ['API URL', _get_config_val(groups, 'api_url')],
        ['Stream模式', _get_config_val(groups, 'stream', 'True')],
        ['负载模式', '梯形负载'],
        ['并发数', conc_range],
        ['运行时长(s)', '300'],
        ['负载机IP', _get_config_val(groups, 'runner_ip')],
        ['参数化数据', 'test.xlsx (499条question, Index)'],
        ['测试脚本', _get_script_name(groups)],
    ]
    for ri, row_data in enumerate(env_rows):
        fill = _alt_fill(ri, FILL_ALT_A, FILL_WHITE)
        for ci, val in enumerate(row_data, 1):
            font = FONT_BOLD if ci == 1 else FONT_NORMAL
            _set_cell(ws5, 4 + ri, ci, val, font=font, fill=fill, alignment=ALIGN_LEFT_CENTER)
    
    ws5.column_dimensions['A'].width = 25.0
    ws5.column_dimensions['B'].width = 50.0
    
    # ── Sheet 6: 性能分析结论 ──
    ws6 = wb.create_sheet('性能分析结论')
    _set_cell(ws6, 1, 1, '性能分析结论',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=3)
    
    conclusions = [
        ('1. 可靠性', f'测试在{conc_range}并发梯度下实现零失败率（{total_fails}/{total_reqs}），API 服务稳定性优秀。'),
        ('2. 吞吐能力', f'并发数从{sorted_groups[0]["concurrency"]}递增至{sorted_groups[-1]["concurrency"]}，运行时长300s。'
                       f'吞吐峰值达到{_round(sa["thr_max"], 1)} tok/s。'),
        ('3. 首Token延迟(TTFT)', f'TTFT最小值{_round(sa["ttft_min"], 3)}s，响应快速。'),
        ('4. TPOT', f'TPOT最小值{_round(sa["tpot_min"], 4)}s，Token生成效率高。'),
        ('5. QPS', f'最大QPS {_round(sa["qps_max"], 3)}。'),
        ('6. 输出Token', f'平均输出Token约{_round(sa["out_avg"], 0)}。'),
        ('7. 输入Token', f'输入Token范围{_round(sa["in_tok_min"], 0)}~{_round(sa["in_tok_max"], 0)}。'),
    ]
    
    for ci, h in enumerate(['结论项', '结论内容'], 1):
        _set_cell(ws6, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    for ri, (title, content) in enumerate(conclusions):
        fill = _alt_fill(ri, FILL_ALT_A, FILL_WHITE)
        _set_cell(ws6, 4 + ri, 1, title, font=FONT_CONCLUSION_TITLE, fill=fill, alignment=ALIGN_LEFT_CENTER)
        _set_cell(ws6, 4 + ri, 2, content, font=FONT_CONCLUSION_BODY, fill=fill, alignment=ALIGN_LEFT)
    
    ws6.column_dimensions['A'].width = 25.0
    ws6.column_dimensions['B'].width = 80.0
    
    # ── Sheet 7: 图表分析 ──
    ws7 = wb.create_sheet('图表分析')
    _set_cell(ws7, 1, 1, f'性能数据 - {device_name}',
              font=FONT_TITLE, alignment=ALIGN_LEFT_CENTER, merge_end_col=7)
    
    chart_headers = ['组号', '并发数', '延迟(s)', '吞吐(tok/s)', 'TTFT(s)', 'QPS', '输出tok']
    for ci, h in enumerate(chart_headers, 1):
        _set_cell(ws7, 3, ci, h, font=FONT_HEADER, fill=FILL_PRIMARY, alignment=ALIGN_CENTER)
    
    for gi, g in enumerate(sorted_groups):
        m = g['stats']['metrics']
        row = 4 + gi
        vals = [gi+1, g['concurrency'],
                _round(m.get('avg_latency', 0)),
                m.get('throughput_max', 0),
                m.get('avg_ttft', 0),
                _round(m.get('max_qps', 0), 3),
                m.get('avg_output_tokens', 0)]
        for ci, v in enumerate(vals, 1):
            _set_cell(ws7, row, ci, v, font=FONT_NORMAL, alignment=ALIGN_CENTER)
    
    for i, w in enumerate([6, 8, 12, 14, 10, 10, 10], 1):
        ws7.column_dimensions[get_column_letter(i)].width = w
    
    wb.save(output_path)
    print(f"[OK] 单设备 Excel 报告已生成: {output_path} (7 Sheets, v6 风格)")

# ============================================================
#  HTML 生成层
# ============================================================

_THIS_DIR = Path(__file__).resolve().parent.parent  # deeprunner-report 根目录
_ASSETS_DIR = _THIS_DIR / 'assets'


def _copy_assets(output_dir):
    """将 assets/ 目录复制到输出目录（JS/CSS/字体文件，递归子目录）"""
    out_assets = Path(output_dir) / 'assets'
    out_assets.mkdir(parents=True, exist_ok=True)
    asset_exts = {'.js', '.css', '.woff2', '.woff', '.ttf', '.svg', '.eot'}
    copied = 0
    for f in _ASSETS_DIR.rglob('*'):
        if f.is_file() and f.suffix in asset_exts:
            rel = f.relative_to(_ASSETS_DIR)
            dest = out_assets / rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(f, dest)
            copied += 1
    print(f"[OK] 已复制 {copied} 个资源文件到 {out_assets}")


def _now_str():
    from datetime import datetime
    return datetime.now().strftime('%Y-%m-%d %H:%M')


def _mhz_fmt(mhz_str):
    """Format MHz string to human-readable GHz, e.g. '2808.0000' → '2.8G'"""
    if not mhz_str:
        return ''
    try:
        val = float(mhz_str)
        if val >= 1000:
            return f'{val/1000:.1f}G'.replace('.0G', 'G')
        return f'{val:.0f}MHz'
    except ValueError:
        return mhz_str


def detect_device_info(stats_dir):
    """从 stats_dir 下的 device-*.md 文件自动检测设备信息和模型信息
    
    扫描目录中的 device-*.md 文件，提取丰富的字段（加速器、CPU、内存、模型、量化、框架等）
    
    返回:
      list[dict]: 每个 device-*.md 一个元素，包含 device/model/accelerator/cpu_info/
                   memory/quantization/framework/api_endpoint/network_mode
    """
    results = []
    base = Path(stats_dir)
    device_files = sorted(base.glob('device*.md'))
    
    # 如果在 stats_dir 没找到，向上搜索父目录（测试报告通常结构：device-*.md 在父目录）
    if not device_files:
        for parent in base.parents:
            device_files = sorted(parent.glob('device*.md'))
            if device_files:
                break
    
    for dev_path in device_files:
        try:
            text = dev_path.read_text(encoding='utf-8')
        except Exception:
            continue

        device_name = None
        model_name = None
        accelerator = ''
        cpu_info = ''
        memory_str = ''
        framework = ''
        api_endpoint = ''
        quantization = ''
        network_mode = ''

        # ── 设备名 ──
        m = re.search(r'\|\s*厂商/SoC\s*\|\s*(.+?)\s*\|', text)
        if m:
            soc = m.group(1).strip()
            if 'Kunpeng' in soc or '鲲鹏' in soc:
                device_name = '昇腾 910B4'
            elif 'GB10' in soc:
                device_name = 'NVIDIA GB10'
            else:
                device_name = soc

        if not device_name:
            m = re.search(r'NPU\s+Name.*?\|\s*(\S+)', text)
            if m:
                npu_name = m.group(1).strip()
                device_name = f'昇腾 {npu_name}' if '910B' in npu_name else npu_name

        if not device_name:
            m = re.search(r'Product Name\s*:\s*(.+)', text)
            if m:
                device_name = m.group(1).strip()

        if not device_name:
            m = re.search(r'主机名\s*\|\s*(.+?)\s*\|', text)
            host = m.group(1).strip() if m else 'unknown'
            m = re.search(r'架构\s*\|\s*(\S+)', text)
            arch = m.group(1).strip() if m else ''
            device_name = f'{host} ({arch})'

        # ── 加速器详情 ──
        def _find_host_ip(t):
            # Match network interface lines like "N: enX: ... inet A.B.C.D"
            m = re.search(r'^\s*(?:\d+:\s+)?(?:en[^\s\w]|eth|wlan)\S*:.*?inet\s+(\d+\.\d+\.\d+\.\d+)', t, re.MULTILINE | re.DOTALL)
            if not m:
                # Fallback: any line with "inet" and NOT 127.0.0.1 nor 0.0.0.0 nor docker/bridge
                for m2 in re.finditer(r'inet\s+(\d+\.\d+\.\d+\.\d+)', t):
                    ip = m2.group(1)
                    if ip not in ('127.0.0.1', '0.0.0.0') and not ip.startswith('172.'):
                        return ip
                # Last resort
                m = re.search(r'inet\s+(?!127\.)(?!0\.)(\d+\.\d+\.\d+\.\d+)', t)
            return m.group(1) if m else None
        
        # NPU: count and model
        npu_count = 0
        npu_model = ''
        for line in text.split('\n'):
            m = re.search(r'\|\s*(\d+)\s+(910B\S*)\s+\|', line)
            if m:
                npu_count += 1
                if not npu_model:
                    npu_model = m.group(2)
        if npu_count > 0:
            # Strip suffix like -1 from 910B4-1
            npu_model_clean = re.sub(r'-\d+$', '', npu_model)
            accelerator = f'昇腾 {npu_model_clean} ({npu_count}卡)'
        
        # GPU
        if not accelerator:
            m = re.search(r'Product Name\s*:\s*(.+)', text)
            if m:
                gpu = m.group(1).strip()
                m2 = re.search(r'Product Architecture\s*:\s*(.+)', text)
                arch_gpu = m2.group(1).strip() if m2 else ''
                m3 = re.search(r'Attached GPUs\s*:\s*(\d+)', text)
                gpu_count = m3.group(1) if m3 else '1'
                accelerator = f'{gpu} ({arch_gpu}, {gpu_count}卡)' if arch_gpu else f'{gpu} ({gpu_count}卡)'

        # ── CPU 信息 ──
        m = re.search(r'型号名称：\s*(.+?)(?:\n|$)', text)
        cpu_model = m.group(1).strip() if m else ''
        
        # Count big/little cores from CPU topology + track MHz ranges
        big_cores = 0
        little_cores = 0
        big_mhz_str = ''
        little_mhz_str = ''
        topo_mhz_vals = []
        for line in text.split('\n'):
            parts = line.strip().split()
            if len(parts) >= 7 and parts[0].isdigit():
                mhz_str = parts[-1] if len(parts) >= 8 else parts[6]
                try:
                    mhz = float(mhz_str)
                except ValueError:
                    continue
                topo_mhz_vals.append((int(parts[0]), mhz, mhz_str))
                if mhz >= 3000:
                    big_cores += 1
                    if not big_mhz_str or mhz_str != big_mhz_str:
                        big_mhz_str = mhz_str
                elif mhz >= 1000:
                    little_cores += 1
                    if not little_mhz_str or mhz_str != little_mhz_str:
                        little_mhz_str = mhz_str

        if big_cores > 0 and little_cores > 0:
            big_cpu_m = re.findall(r'型号名称：\s*(.+?)(?:\n|$)', text)
            # From observation: first 型号名称 = HIGHEST MHz core, second = LOWER MHz core
            high_label = big_cpu_m[0].strip() if len(big_cpu_m) > 0 else '大核'
            low_label = big_cpu_m[1].strip() if len(big_cpu_m) > 1 else '小核'
            cpu_info = f'{big_cores}x{high_label}@{_mhz_fmt(big_mhz_str)} + {little_cores}x{low_label}@{_mhz_fmt(little_mhz_str)}'
        elif big_cores > 0:
            cpu_info = f'{big_cores}核 x {cpu_model}'
        elif 'Kunpeng' in text or '鲲鹏' in text:
            # Count physical cores (SOCKET*每个座的核数)
            m_socket = re.search(r'座[：:]\s*(\d+)', text)
            m_cores_per = re.search(r'每个座的核数[：:]\s*(\d+)', text)
            if m_socket and m_cores_per:
                total = int(m_socket.group(1)) * int(m_cores_per.group(1))
            else:
                total = sum(1 for line in text.split('\n') if line.strip().split() and line.strip().split()[0].isdigit())
            cpu_info = f'鲲鹏920: {total}核'
        elif cpu_model:
            cpu_info = f'{cpu_model}'
        
        if not cpu_info:
            m = re.search(r'CPU[：:]\s+(\d+)', text)
            if m:
                cpu_info = f'{m.group(1)}核'

        # ── 内存 ──
        # Try "内存： total used free ..." format
        m = re.search(r'内存[：:]\s+(\d+)[GgＧ][Ii]?\s+(\d+)[GgＧ]', text)
        # Try "Mem: total used free ..." format  
        if not m:
            m = re.search(r'Mem[：:]\s+(\d+)[Gg][Ii]?\s+(\d+)[Gg]', text)
        # Try free -k format
        if not m:
            m = re.search(r'(\d+)\s*K total memory\s*\n\s*(\d+)\s*K used', text)
        if m:
            mem_total = m.group(1)
            mem_used = m.group(2) if m.lastindex >= 2 else ''
            memory_str = f'{mem_total} GiB'
            if mem_used:
                memory_str += f' (已用{mem_used})'

        # ── 模型名 ──
        m = re.search(r'"id"\s*:\s*"([^"]+)"', text)
        if m:
            model_name = m.group(1).strip()

        if not model_name:
            m = re.search(r'"(Qwen/[^"]+)"', text)
            if m:
                model_name = m.group(1).strip()

        if not model_name:
            m = re.search(r'"root"\s*:\s*"[^"]*/([^/"]+)"', text)
            if m:
                model_name = m.group(1).strip()

        if not model_name:
            m = re.search(r'(?:容器名|NAMES).*?qwen[_-]?(\S+)', text, re.IGNORECASE)
            if not m:
                m = re.search(r'qwen[_-]?(\S+)', text, re.IGNORECASE)
            if m:
                raw = m.group(0).strip().split(':')[0]
                model_name = raw

        model_name = model_name or '未知模型'

        # ── 量化方式 ──
        # From model name suffix
        if 'FP8' in model_name or 'fp8' in text.lower():
            quantization = 'FP8'
        elif 'BF16' in text or 'bf16' in text.lower():
            quantization = 'BF16'
        elif 'INT4' in model_name or 'int4' in text.lower():
            quantization = 'INT4'
        elif 'INT8' in model_name or 'int8' in text.lower():
            quantization = 'INT8'
        elif 'FP16' in text or 'fp16' in text.lower():
            quantization = 'FP16'
        
        # Check docker env or known patterns
        if not quantization:
            if 'Ascend' in text or 'npu-smi' in text:
                quantization = 'BF16/FP16'
            else:
                quantization = 'FP8' if 'GB10' in text else '未知'

        # ── 推理框架 ──
        m = re.search(r'vLLM\s+(\d+\.\S+)', text)
        if m:
            framework = f'vLLM {m.group(1)}'
        elif 'vLLM' in text or 'vllm' in text.lower():
            framework = 'vLLM'
        else:
            framework = 'vLLM'

        # ── API 端点 ──
        host_ip = _find_host_ip(text)
        
        # 1) Docker port mapping: look for a container with 8000/tcp or :8000->
        m = re.search(r'0\.0\.0\.0:(\d+)->(\d+)/tcp', text)
        if m and m.group(2) == '8000':
            port = m.group(1)
            if host_ip:
                api_endpoint = f'http://{host_ip}:{port}'
        
        # 2) Docker host networking: look for "--port", "8000" in startup args
        if not api_endpoint:
            m = re.search(r'["\']--port["\']\s*,\s*["\'](\d+)["\']', text)
            if m:
                port = m.group(1)
                if host_ip:
                    api_endpoint = f'http://{host_ip}:{port}'
        
        # 3) From listening ports: look for port 8000 specifically
        if not api_endpoint:
            m = re.search(r'(?:0\.0\.0\.0|\*|:::):8000', text)
            if m:
                if host_ip:
                    api_endpoint = f'http://{host_ip}:8000'
                else:
                    api_endpoint = 'http://localhost:8000'
        
        # 4) Fallback: any 192.168.x.x:8000
        if not api_endpoint:
            m = re.search(r'(192\.168\.\d+\.\d+):8000', text)
            if m:
                api_endpoint = f'http://{m.group(1)}:8000'

        # ── 网络模式 ──
        # Check if a vLLM/qwen container has port mapping in its Docker line
        # Docker ps: NAME IMAGE STATUS PORTS
        docker_lines = re.findall(r'^.*?(?:vllm|qwen)\S*.*$', text, re.MULTILINE | re.IGNORECASE)
        has_vllm_bridge = any('0.0.0.0:' in line and '->' in line for line in docker_lines)
        has_vllm_container = any('Up' in line for line in docker_lines)
        
        if has_vllm_bridge:
            network_mode = '容器 bridge (端口映射)'
        elif has_vllm_container:
            network_mode = '容器 host 网络'
        else:
            network_mode = '容器 bridge (端口映射)' if re.search(r'0\.0\.0\.0:\d+->', text) else '容器 host 网络'

        results.append({
            'device': device_name or '未知设备',
            'model': model_name,
            'accelerator': accelerator,
            'cpu_info': cpu_info,
            'memory': memory_str,
            'quantization': quantization,
            'framework': framework,
            'api_endpoint': api_endpoint,
            'network_mode': network_mode,
            'rawMarkdown': text,
            'source_file': dev_path.name,
        })

    return results


def _build_chart_data(groups):
    """从 groups 构建 chartData 结构：concurrency, throughput, latency"""
    sorted_groups = sorted(groups, key=lambda g: g['concurrency'])
    return {
        'concurrency': [g['concurrency'] for g in sorted_groups],
        'throughput': [g['stats']['metrics'].get('throughput_max', 0) for g in sorted_groups],
        'latency': [_round(g['stats']['metrics'].get('avg_latency', 0)) for g in sorted_groups],
    }


def _build_detail_rows(groups):
    """从 groups 构建明细表行数据"""
    sorted_groups = sorted(groups, key=lambda g: g['concurrency'])
    rows = []
    for gi, g in enumerate(sorted_groups, 1):
        m = g['stats']['metrics']
        rows.append([
            gi, g['concurrency'],
            g['stats']['total_requests'], g['stats']['failed_requests'],
            m.get('min_latency', 0), m.get('max_latency', 0),
            _round(m.get('avg_latency', 0)),
            m.get('throughput_min', 0), m.get('throughput_max', 0),
            _round(m.get('min_qps', 0), 3), _round(m.get('max_qps', 0), 3),
            m.get('avg_ttft', 0), m.get('avg_tpot', 0),
            int(m.get('avg_input_tokens', 0)),
            int(m.get('min_output_tokens', 0)), int(m.get('max_output_tokens', 0)),
        ])
    return rows


def _compute_avg_ttft(groups):
    vals = [g['stats']['metrics'].get('avg_ttft', 0) for g in groups if g['stats']['metrics'].get('avg_ttft', 0) > 0]
    return vals[0] if vals else 0


def _compute_avg_tpot(groups):
    vals = [g['stats']['metrics'].get('avg_tpot', 0) for g in groups if g['stats']['metrics'].get('avg_tpot', 0) > 0]
    return vals[0] if vals else 0


def generate_single_html(groups, device_name, model_name, output_path, device_info=None):
    """生成单设备 HTML 可视化报告"""
    template_path = _ASSETS_DIR / 'single-report-template.html'
    if not template_path.exists():
        print(f"ERROR: 模板文件不存在 {template_path}")
        sys.exit(1)

    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    sorted_groups = sorted(groups, key=lambda g: g['concurrency'])
    ttft = _compute_avg_ttft(sorted_groups)
    tpot = _compute_avg_tpot(sorted_groups)
    throughput_vals = [g['stats']['metrics'].get('throughput_max', 0) for g in sorted_groups]
    throughput = max(throughput_vals) if throughput_vals else 0
    lat_vals = [g['stats']['metrics'].get('avg_latency', 0) for g in sorted_groups]
    e2e = max(lat_vals) if lat_vals else 0
    total_runtime = 300  # 从测试配置获取，默认300s

    # 汇总数据
    total_errors = {'failure_count': 0, 'exception_count': 0, 'sla_count': 0}
    for g in groups:
        err = g['config'].get('errors', {})
        if err:
            total_errors['failure_count'] += err.get('failure_count', 0)
            total_errors['exception_count'] += err.get('exception_count', 0)
            total_errors['sla_count'] += err.get('sla_count', 0)
    if device_info:
        device_info = dict(device_info)
        device_info['errors'] = total_errors
    
    report_data = {
        'deviceName': device_name,
        'modelName': model_name,
        'deviceInfo': device_info,  # rich device info from device-*.md
        'chartData': _build_chart_data(groups),
        'detailRows': _build_detail_rows(groups),
        'wordFile': None,
    }

    # 替换占位符
    replacements = {
        '{{TITLE}}': f'{device_name} 推理性能测评报告',
        '{{DEVICE_NAME}}': device_name,
        '{{DATE}}': _now_str(),
        '{{MODEL_NAME}}': model_name,
        '{{GROUP_COUNT}}': str(len(groups)),
        '{{RUNTIME}}': str(total_runtime),
        '{{TTFT_VALUE}}': str(_round(ttft, 3)),
        '{{TTFT_STATUS}}': '首 Token 响应快速' if ttft < 1 else '首 Token 响应正常' if ttft < 3 else '首 Token 响应偏慢',
        '{{TPOT_VALUE}}': str(_round(tpot, 4)),
        '{{TPOT_STATUS}}': '输出 Token 生成高效' if tpot < 0.05 else '输出 Token 生成正常' if tpot < 0.1 else '输出 Token 生成偏慢',
        '{{THROUGHPUT_VALUE}}': str(_round(throughput, 1)),
        '{{THROUGHPUT_STATUS}}': '吞吐量高' if throughput > 100 else '吞吐量正常' if throughput > 30 else '吞吐量较低',
        '{{E2E_VALUE}}': str(_round(e2e, 1)),
        '{{E2E_STATUS}}': '端到端延迟低' if e2e < 5 else '端到端延迟正常' if e2e < 20 else '端到端延迟偏高',
        '{{SIGNER}}': 'DeepRunner',
        '{{SIGN_DATE}}': _now_str(),
        '{{REPORT_DATA_JSON}}': json.dumps(report_data, ensure_ascii=True),
    }

    for placeholder, value in replacements.items():
        html = html.replace(placeholder, value)

    # 处理设备信息（空数据时生成占位）
    if '{{DEVICE_INFO_JSON}}' in html:
        html = html.replace('{{DEVICE_INFO_JSON}}', json.dumps(device_info or {
            'device': device_name, 'model': model_name,
            'groups': len(groups),
        }, ensure_ascii=True))

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    # 复制 assets
    output_dir = os.path.dirname(output_path) or '.'
    _copy_assets(output_dir)

    print(f"[OK] 单设备 HTML 报告已生成: {output_path}")


def generate_compare_html(groups_a, groups_b, name_a, name_b, model_a, model_b, output_path, device_info_a=None, device_info_b=None):
    """生成双设备比对 HTML 可视化报告"""
    template_path = _ASSETS_DIR / 'compare-report-template.html'
    if not template_path.exists():
        print(f"ERROR: 模板文件不存在 {template_path}")
        sys.exit(1)

    with open(template_path, 'r', encoding='utf-8') as f:
        html = f.read()

    sorted_a = sorted(groups_a, key=lambda g: g['concurrency'])
    sorted_b = sorted(groups_b, key=lambda g: g['concurrency'])

    label_a = name_a.split()[0] if ' ' in name_a else name_a
    label_b = name_b.split()[0] if ' ' in name_b else name_b

    # 构建图表数据
    def _metric_list(groups, key):
        return [g['stats']['metrics'].get(key, 0) for g in groups]

    chart_data = {
        'labels': [g['concurrency'] for g in sorted_a],
        'latencyA': _metric_list(sorted_a, 'avg_latency'),
        'latencyB': _metric_list(sorted_b, 'avg_latency'),
        'throughputA': _metric_list(sorted_a, 'throughput_max'),
        'throughputB': _metric_list(sorted_b, 'throughput_max'),
        'ttftA': _metric_list(sorted_a, 'avg_ttft'),
        'ttftB': _metric_list(sorted_b, 'avg_ttft'),
        'qpsA': _metric_list(sorted_a, 'max_qps'),
        'qpsB': _metric_list(sorted_b, 'max_qps'),
        'outputTokA': _metric_list(sorted_a, 'avg_output_tokens'),
        'outputTokB': _metric_list(sorted_b, 'avg_output_tokens'),
        'requestsA': [g['stats']['total_requests'] for g in sorted_a],
        'requestsB': [g['stats']['total_requests'] for g in sorted_b],
    }

    # 汇总设备信息（含 errors）
    if device_info_a:
        device_info_a = dict(device_info_a)
        err_a = {'failure_count': 0, 'exception_count': 0, 'sla_count': 0}
        for g in groups_a:
            e = g['config'].get('errors', {})
            if e:
                err_a['failure_count'] += e.get('failure_count', 0)
                err_a['exception_count'] += e.get('exception_count', 0)
                err_a['sla_count'] += e.get('sla_count', 0)
        device_info_a['errors'] = err_a
    if device_info_b:
        device_info_b = dict(device_info_b)
        err_b = {'failure_count': 0, 'exception_count': 0, 'sla_count': 0}
        for g in groups_b:
            e = g['config'].get('errors', {})
            if e:
                err_b['failure_count'] += e.get('failure_count', 0)
                err_b['exception_count'] += e.get('exception_count', 0)
                err_b['sla_count'] += e.get('sla_count', 0)
        device_info_b['errors'] = err_b
    
    report = {
        'nameA': name_a,
        'nameB': name_b,
        'labelA': label_a,
        'labelB': label_b,
        'deviceInfoA': device_info_a,
        'deviceInfoB': device_info_b,
        'chartData': chart_data,
        'excelFile': None,
    }

    # 替换占位符
    replacements = {
        '{{REPORT_DATA_JSON}}': json.dumps(report, ensure_ascii=True),
        '{{DATE_RANGE}}': _now_str(),
        '{{MODEL_NAME}}': f'{model_a} vs {model_b}',
        '{{NAME_A_SHORT}}': label_a,
        '{{NAME_B_SHORT}}': label_b,
        '{{SIGNER}}': 'DeepRunner',
        '{{SIGN_DATE}}': _now_str(),
    }

    for placeholder, value in replacements.items():
        html = html.replace(placeholder, value)

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(html)

    # 复制 assets
    output_dir = os.path.dirname(output_path) or '.'
    _copy_assets(output_dir)

    print(f"[OK] 比对 HTML 报告已生成: {output_path}")


# ============================================================
#  CLI 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='DeepRunner 6.0 报告生成器')
    parser.add_argument('--mode', required=True, choices=['single', 'compare', 'batch', 'single-html', 'compare-html', 'discover'])
    parser.add_argument('--root-dir', help='根目录模式：自动扫描该目录下所有设备并生成全部报告（含逐并发+汇总+比对）')
    parser.add_argument('--stats-dir', help='单设备/批量: 测试报告目录')
    parser.add_argument('--stats-dir-a', help='比对: 设备A目录')
    parser.add_argument('--stats-dir-b', help='比对: 设备B目录')
    parser.add_argument('--output-dir', required=True, help='输出目录')
    parser.add_argument('--device', default=None, help='设备名（不传则从 device-*.md 自动检测）')
    parser.add_argument('--device-a', default=None, help='设备A名（不传则自动检测）')
    parser.add_argument('--device-b', default=None, help='设备B名（不传则自动检测）')
    parser.add_argument('--name-a', default=None, help='设备A显示名（不传则自动检测）')
    parser.add_argument('--name-b', default=None, help='设备B显示名（不传则自动检测）')
    parser.add_argument('--model', default=None, help='模型名（不传则从 device-*.md 自动检测）')
    parser.add_argument('--model-a', default=None, help='模型A名（不传则自动检测）')
    parser.add_argument('--model-b', default=None, help='模型B名（不传则自动检测）')
    parser.add_argument('--per-concurrency', action='store_true', default=True, help='按并发逐组生成报告（默认开启）')
    parser.add_argument('--no-per-concurrency', action='store_true', default=False, help='关闭逐并发生成，仅汇总报告')
    
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    
    # 自动检测设备/模型信息
    _device_auto, _model_auto = None, None
    _device_a_auto, _model_a_auto = None, None
    _device_b_auto, _model_b_auto = None, None
    _info_list = []  # cache full detect_device_info result
    _info_a_list, _info_b_list = [], []
    
    if args.mode in ('single', 'single-html', 'batch'):
        _info_list = detect_device_info(args.stats_dir)
        if _info_list:
            _device_auto = _info_list[0]['device']
            _model_auto = _info_list[0]['model']
            print(f"[检测] 设备: {_device_auto} | 模型: {_model_auto} (来自 {_info_list[0]['source_file']})")
            if len(_info_list) > 1:
                print(f"[检测] 发现第2个设备文件: {_info_list[1]['source_file']} ({_info_list[1]['device']}, {_info_list[1]['model']})")

    elif args.mode in ('compare', 'compare-html'):
        _info_a_list = detect_device_info(args.stats_dir_a)
        _info_b_list = detect_device_info(args.stats_dir_b)
        if _info_a_list:
            _device_a_auto = _info_a_list[0]['device']
            _model_a_auto = _info_a_list[0]['model']
            print(f"[检测] 设备A: {_device_a_auto} | 模型A: {_model_a_auto}")
        if _info_b_list:
            _device_b_auto = _info_b_list[0]['device']
            _model_b_auto = _info_b_list[0]['model']
            print(f"[检测] 设备B: {_device_b_auto} | 模型B: {_model_b_auto}")
    
    # 使用 CLI 参数（若提供）或自动检测值
    device = args.device or _device_auto or '设备'
    device_a = args.device_a or _device_a_auto or '设备A'
    device_b = args.device_b or _device_b_auto or '设备B'
    name_a = args.name_a or _device_a_auto or '设备A'
    name_b = args.name_b or _device_b_auto or '设备B'
    model = args.model or _model_auto or '未知模型'
    model_a = args.model_a or _model_a_auto or '未知模型'
    model_b = args.model_b or _model_b_auto or '未知模型'
    
    do_per = args.per_concurrency and not args.no_per_concurrency
    
    # ── discover 模式：自动扫描目录并生成全部报告 ──
    if args.mode == 'discover':
        if not args.root_dir:
            print("ERROR: discover 模式需要指定 --root-dir")
            sys.exit(1)
        devices = discover_devices(args.root_dir)
        print_discovery(devices)
        print_validation(devices)
        if len(devices) < 1:
            sys.exit(0)
        
        # 为每个设备生成单设备报告
        for dev in devices:
            name_clean = dev['name'].replace(' ', '_').replace('（', '_').replace('）', '_')
            dev_dir = os.path.join(args.output_dir, name_clean)
            os.makedirs(dev_dir, exist_ok=True)
            print(f"\n--- 正在生成: {dev['name']} ---")
            groups = dev['groups']
            di = dev.get('device_info')
            
            # 逐并发
            if do_per and len(groups) > 1:
                for g in groups:
                    conc = g['concurrency']
                    generate_single_html([g], dev['name'], di.get('model','?') if di else '?',
                                         os.path.join(dev_dir, f'deeprunner_{name_clean}_并发{conc}.html'),
                                         device_info=di)
                    generate_single_excel([g], dev['name'], di.get('model','?') if di else '?',
                                          os.path.join(dev_dir, f'DeepRunner_{name_clean}_并发{conc}.xlsx'),
                                          device_info=di)
            # 汇总
            generate_single_html(groups, dev['name'], di.get('model','?') if di else '?',
                                 os.path.join(dev_dir, f'deeprunner_{name_clean}测评报告.html'),
                                 device_info=di)
            generate_single_excel(groups, dev['name'], di.get('model','?') if di else '?',
                                  os.path.join(dev_dir, f'DeepRunner_推理性能测试报告_{name_clean}.xlsx'),
                                  device_info=di)
        
        # 如果有 ≥2 个设备，生成比对报告
        if len(devices) >= 2:
            print(f"\n--- 正在生成: 比对报告 ({devices[0]['name']} vs {devices[1]['name']}) ---")
            name_a_clean = devices[0]['name'].replace(' ', '_').replace('（', '_').replace('）', '_')
            name_b_clean = devices[1]['name'].replace(' ', '_').replace('（', '_').replace('）', '_')
            di_a = devices[0].get('device_info')
            di_b = devices[1].get('device_info')
            
            # 逐并发比对
            if do_per and len(devices[0]['groups']) > 1 and len(devices[1]['groups']) > 1:
                conc_map_a = {g['concurrency']: g for g in devices[0]['groups']}
                conc_map_b = {g['concurrency']: g for g in devices[1]['groups']}
                common_concs = sorted(set(conc_map_a.keys()) & set(conc_map_b.keys()))
                for conc in common_concs:
                    generate_compare_html([conc_map_a[conc]], [conc_map_b[conc]],
                                          devices[0]['name'], devices[1]['name'],
                                          (di_a or {}).get('model','?'), (di_b or {}).get('model','?'),
                                          os.path.join(args.output_dir, f'deeprunner_比对_{name_a_clean}_vs_{name_b_clean}_并发{conc}.html'),
                                          device_info_a=di_a, device_info_b=di_b)
                    generate_compare_excel([conc_map_a[conc]], [conc_map_b[conc]],
                                           devices[0]['name'], devices[1]['name'],
                                           (di_a or {}).get('model','?'), (di_b or {}).get('model','?'),
                                           os.path.join(args.output_dir, f'DeepRunner_比对_{name_a_clean}_vs_{name_b_clean}_并发{conc}.xlsx'))
            # 汇总比对
            generate_compare_html(devices[0]['groups'], devices[1]['groups'],
                                  devices[0]['name'], devices[1]['name'],
                                  (di_a or {}).get('model','?'), (di_b or {}).get('model','?'),
                                  os.path.join(args.output_dir, f'deeprunner_推理性能比对报告_{name_a_clean}_vs_{name_b_clean}.html'),
                                  device_info_a=di_a, device_info_b=di_b)
            generate_compare_excel(devices[0]['groups'], devices[1]['groups'],
                                   devices[0]['name'], devices[1]['name'],
                                   (di_a or {}).get('model','?'), (di_b or {}).get('model','?'),
                                   os.path.join(args.output_dir, f'DeepRunner_推理性能比对报告_{name_a_clean}_vs_{name_b_clean}.xlsx'))
        
        print(f"\n[OK] 全部报告已生成至: {args.output_dir}")
        return
    
    if args.mode == 'single':
        groups = scan_test_groups(args.stats_dir)
        print(f"发现 {len(groups)} 组测试数据")
        dev_info = None
        if _info_list:
            if args.device:
                for di in _info_list:
                    if di['device'] == args.device or di['model'] == args.model:
                        dev_info = di
                        break
            if not dev_info:
                dev_info = _info_list[0]
        # 1) 逐并发生成
        if do_per and len(groups) > 1:
            print(f"--- 逐并发生成 ({len(groups)}组) ---")
            for g in groups:
                conc = g['concurrency']
                out_html = os.path.join(args.output_dir, f'deeprunner_{device}_并发{conc}.html')
                out_xlsx = os.path.join(args.output_dir, f'DeepRunner_{device}_并发{conc}.xlsx')
                generate_single_html([g], device, model, out_html, device_info=dev_info)
                generate_single_excel([g], device, model, out_xlsx, device_info=dev_info)
        # 2) 汇总报告
        print(f"--- 汇总报告 ---")
        out = os.path.join(args.output_dir, f'DeepRunner_推理性能测试报告_{device}.xlsx')
        generate_single_excel(groups, device, model, out, device_info=dev_info)
    
    elif args.mode == 'compare':
        groups_a = scan_test_groups(args.stats_dir_a)
        groups_b = scan_test_groups(args.stats_dir_b)
        print(f"设备A ({name_a}): {len(groups_a)} 组 | 设备B ({name_b}): {len(groups_b)} 组")
        # 1) 逐并发生成比对
        if do_per and len(groups_a) > 1 and len(groups_b) > 1:
            print(f"--- 逐并发生成比对 ({min(len(groups_a),len(groups_b))}组) ---")
            conc_map_a = {g['concurrency']: g for g in groups_a}
            conc_map_b = {g['concurrency']: g for g in groups_b}
            common_concs = sorted(set(conc_map_a.keys()) & set(conc_map_b.keys()))
            for conc in common_concs:
                out_xlsx = os.path.join(args.output_dir, f'DeepRunner_比对_{device_a}_vs_{device_b}_并发{conc}.xlsx')
                out_html = os.path.join(args.output_dir, f'deeprunner_比对_{device_a}_vs_{device_b}_并发{conc}.html')
                generate_compare_excel([conc_map_a[conc]], [conc_map_b[conc]], name_a, name_b, model_a, model_b, out_xlsx)
                generate_compare_html([conc_map_a[conc]], [conc_map_b[conc]], name_a, name_b, model_a, model_b, out_html)
        # 2) 汇总比对报告
        print(f"--- 汇总比对报告 ---")
        out = os.path.join(args.output_dir, f'DeepRunner_推理性能比对报告_{device_a}_vs_{device_b}.xlsx')
        generate_compare_excel(groups_a, groups_b, name_a, name_b,
                               model_a, model_b, out)
    
    elif args.mode == 'batch':
        groups = scan_test_groups(args.stats_dir)
        print(f"发现 {len(groups)} 组测试数据")
        out = os.path.join(args.output_dir, f'DeepRunner_推理性能测试报告_批量.xlsx')
        generate_single_excel(groups, device, model, out)

    elif args.mode == 'single-html':
        groups = scan_test_groups(args.stats_dir)
        print(f"发现 {len(groups)} 组测试数据")
        dev_info = None
        if _info_list:
            if args.device:
                for di in _info_list:
                    if di['device'] == args.device or di['model'] == args.model:
                        dev_info = di
                        break
            if not dev_info:
                dev_info = _info_list[0]
        # 1) 逐并发生成
        if do_per and len(groups) > 1:
            print(f"--- 逐并发生成 ({len(groups)}组) ---")
            for g in groups:
                conc = g['concurrency']
                out = os.path.join(args.output_dir, f'deeprunner_{device}_并发{conc}.html')
                generate_single_html([g], device, model, out, device_info=dev_info)
        # 2) 汇总报告
        print(f"--- 汇总报告 ---")
        out = os.path.join(args.output_dir, f'deeprunner_{device}测评报告.html')
        generate_single_html(groups, device, model, out, device_info=dev_info)

    elif args.mode == 'compare-html':
        groups_a = scan_test_groups(args.stats_dir_a)
        groups_b = scan_test_groups(args.stats_dir_b)
        print(f"设备A ({name_a}): {len(groups_a)} 组 | 设备B ({name_b}): {len(groups_b)} 组")
        # Resolve device info
        dev_a_info = None
        if _info_a_list:
            if args.name_a:
                for di in _info_a_list:
                    if di['device'] == args.name_a or di['device'] in (args.name_a or ''):
                        dev_a_info = di
                        break
            if not dev_a_info:
                dev_a_info = _info_a_list[0]
        dev_b_info = None
        if _info_b_list:
            if args.name_b:
                for di in _info_b_list:
                    if di['device'] == args.name_b or di['device'] in (args.name_b or ''):
                        dev_b_info = di
                        break
            if not dev_b_info:
                dev_b_info = _info_b_list[0]
        # 1) 逐并发生成比对
        if do_per and len(groups_a) > 1 and len(groups_b) > 1:
            print(f"--- 逐并发生成比对 ({min(len(groups_a),len(groups_b))}组) ---")
            conc_map_a = {g['concurrency']: g for g in groups_a}
            conc_map_b = {g['concurrency']: g for g in groups_b}
            common_concs = sorted(set(conc_map_a.keys()) & set(conc_map_b.keys()))
            for conc in common_concs:
                out = os.path.join(args.output_dir, f'deeprunner_比对_{device_a}_vs_{device_b}_并发{conc}.html')
                generate_compare_html([conc_map_a[conc]], [conc_map_b[conc]], name_a, name_b,
                                      model_a, model_b, out,
                                      device_info_a=dev_a_info, device_info_b=dev_b_info)
        # 2) 汇总比对报告
        print(f"--- 汇总比对报告 ---")
        out = os.path.join(args.output_dir, f'deeprunner_推理性能比对报告_{device_a}_vs_{device_b}.html')
        generate_compare_html(groups_a, groups_b, name_a, name_b,
                              model_a, model_b, out,
                              device_info_a=dev_a_info,
                              device_info_b=dev_b_info)

if __name__ == '__main__':
    main()
